from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "Product Description"
ws['A2'] = "Shirt 100% Cotton"
wb.save("test.xlsx")
